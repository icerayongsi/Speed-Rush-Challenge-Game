import sqlite3
import pandas as pd
from datetime import datetime, timedelta
import os
import base64
import tempfile
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import io
import shutil
import pytz
import re

def get_db_path():
    """Get the path to the SQLite database file."""
    script_dir = Path(__file__).parent.absolute()
    return script_dir / 'data' / 'game_data.db'

def format_datetime(dt_str):
    """Format datetime string to local time (GMT+7)."""
    try:
        dt = datetime.strptime(dt_str, '%Y-%m-%d %H:%M:%S')
    except:
        dt = datetime.strptime(dt_str, '%Y-%m-%dT%H:%M:%S.%fZ')
    
    # Convert to GMT+7
    gmt7 = pytz.timezone('Asia/Bangkok')
    utc = pytz.utc
    
    # If datetime is naive, assume it's in UTC
    if dt.tzinfo is None:
        dt = utc.localize(dt)
    
    return dt.astimezone(gmt7).strftime('%Y-%m-%d %H:%M:%S')

def export_game_history():
    """Export game history to an Excel file similar to /api/export-history endpoint."""
    db_path = get_db_path()
    
    # Check if database exists
    if not db_path.exists():
        print(f"Error: Database file not found at {db_path}")
        return
    
    conn = None
    try:
        # Connect to the SQLite database
        conn = sqlite3.connect(str(db_path))
        
        # Query to get all game sessions with user details
        query = """
        SELECT 
            u.name as name,
            g.score as score,
            g.duration as duration,
            g.played_at as played_at,
            u.business_card as business_card
        FROM game_sessions g
        JOIN users u ON g.user_id = u.id
        ORDER BY g.played_at DESC
        """
        
        # Read data into a pandas DataFrame
        df = pd.read_sql_query(query, conn)
        
        if df.empty:
            print("No game data found to export.")
            return
        
        # Create output directory if it doesn't exist
        output_dir = Path(__file__).parent.absolute() / 'exports'
        output_dir.mkdir(exist_ok=True, parents=True)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = output_dir / f'game_history_export_{timestamp}.xlsx'
        
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Game History"
        
        # Define column headers and widths
        headers = [
            ("Name", 30),
            ("Score", 15),
            ("Duration (seconds)", 20),
            ("Played At", 30),
            ("Business Card", 30)
        ]
        
        # Write headers
        for col_num, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Set default row height
        row_height = 25
            
        # Write data rows
        for row_num, (_, row) in enumerate(df.iterrows(), 2):
            # Set row height
            ws.row_dimensions[row_num].height = row_height
            
            # Write regular data
            ws.cell(row=row_num, column=1, value=row['name'])
            ws.cell(row=row_num, column=2, value=row['score'])
            ws.cell(row=row_num, column=3, value=row['duration'])
            ws.cell(row=row_num, column=4, value=format_datetime(row['played_at']))
            
            # Handle business card image
            if row['business_card'] and pd.notna(row['business_card']):
                try:
                    # Get the full path to the image file
                    img_filename = Path(row['business_card']).name
                    img_path = Path(__file__).parent.absolute() / 'data' / 'uploads' / img_filename
                    print(img_path)
                    if img_path.exists() and img_path.is_file():
                        
                        # Load and resize image
                        with PILImage.open(img_path) as img:
                            # Resize maintaining aspect ratio (max 200x100)
                            img.thumbnail((200, 100), PILImage.Resampling.LANCZOS)
                            
                            # Save resized image to bytes
                            img_byte_arr = io.BytesIO()
                            img_format = 'PNG' if img.format == 'PNG' else 'JPEG'
                            img.save(img_byte_arr, format=img_format)
                            img_byte_arr.seek(0)
                            
                            # Create openpyxl image and add to worksheet
                            xl_img = XLImage(io.BytesIO(img_byte_arr.getvalue()))
                            xl_img.width = img.width
                            xl_img.height = img.height
                            
                            # Add image to worksheet (column E)
                            cell_ref = f'E{row_num}'
                            xl_img.anchor = cell_ref
                            ws.add_image(xl_img)
                            
                            # Set row height to fit image (converting pixels to Excel units)
                            ws.row_dimensions[row_num].height = img.height * 0.75
                            
                            # Set column width to fit image (converting pixels to Excel units)
                            ws.column_dimensions['E'].width = max(ws.column_dimensions['E'].width, img.width / 7)
                            
                            # Clean up
                            img_byte_arr.close()
                    else:
                        ws.cell(row=row_num, column=5, value="Image not found")
                        
                except Exception as e:
                    print(f"Error processing image {img_filename}: {e}")
                    ws.cell(row=row_num, column=5, value="Error loading image")
            else:
                ws.cell(row=row_num, column=5, value="No business card")
            
            # Set alignment for all cells
            for col in range(1, 6):
                cell = ws.cell(row=row_num, column=col)
                if col != 5:  # Skip image column
                    cell.alignment = Alignment(
                        horizontal='center', 
                        vertical='center', 
                        wrap_text=True
                    )
                
        # Auto-adjust column widths based on content
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            # Skip image column (E)
            if column_letter == 'E':
                continue
                
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        # Save the workbook
        wb.save(output_file)
        print(f"Export completed successfully. File saved to: {output_file}")
        
    except sqlite3.Error as e:
        print(f"Database error: {e}")
    except Exception as e:
        print(f"Error during export: {e}")
    finally:
        if 'conn' in locals() and conn is not None:
            conn.close()

if __name__ == "__main__":
    export_game_history()
